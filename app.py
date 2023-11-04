from flask import Flask, request, render_template
from openpyxl import load_workbook

app = Flask(__name__)

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/find_replace', methods=['POST'])
def find_replace():
    data = request.get_json()
    mword = data["mword"]
    rword = data["rword"]
    texta = data["texta"]

    # Open the Excel file
    wb = load_workbook('wordslist.xlsx')
    ws = wb['Sheet1']

    # Add the user-provided find and replace pair as a new row in the Excel worksheet
    new_row = (mword, rword)
    ws.append(new_row)

    # Loop through the Excel file and perform find and replace based on its content
    modified_text = texta
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        findWord, replaceWord = row
        modified_text = modified_text.replace(findWord, replaceWord)

    # Save the Excel file with the new row
    wb.save('wordslist.xlsx')

    return modified_text

if __name__ == '__main__':
    app.run(debug=True)